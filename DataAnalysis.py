#to interact with excel
from openpyxl import *
from openpyxl.utils.cell import get_column_letter
import os
#to compute the linear regression
from scipy.stats import linregress

def linear_comparison(sheet, vector, feature, key, n_coloumn):
    """
    The Function computes a linear regression between a vector and all the columns of a dataset.
    Every column of the dataset represents the value of a feature for each considered interview
    :param sheet: dataset of the features
    :param vector: vector in which every element represents the mean of all the answers associated with an interview
    :param feature: the name of the dataset
    :param key: the name of the vector
    :param n_coloumn: the number of the columns of the dataset
    :return: It prints the rho-value and the p-value of all the regressions which have a rho-value> 0.4 or < -0.4
    """
    print("Regressione tra", feature, "e le risposte di", key,":")
    for i in range(2, n_coloumn):  # n_coloumn
        x = []
        list = sheet[get_column_letter(i)]
        for j in range(1, len(list)):
            x.append(list[j].value)
        if (linregress(x, vector)[2] <= (-0.4) or linregress(x, vector)[2] >= 0.4):
            print("Rho con vettore", list[0].value, ":", linregress(x, vector)[2], "p value", linregress(x, vector)[3])
    print()


class DataAnalysis:
    # Normalized average vectors
    a1=[]
    a2=[]
    b1=[]
    b2=[]
    # Normalized average vector with only peer answer
    a1_peer=[]
    a2_peer=[]
    b1_peer=[]
    b2_peer=[]
    # Normalized average vector from selected questions
    a1_selected=[]
    a2_selected=[]
    b1_selected=[]
    b2_selected=[]
    a1_selected_peer=[]
    a2_selected_peer=[]
    b1_selected_peer=[]
    b2_selected_peer=[]
    features_list=[]
    #dictionary which contains the concatenation beetween the vectors
    a1a2dictionary ={}
    b1b2dictionary={}
    alldictionary ={}



    def __init__(self):
        """
        This function initializes the object
        """
        #Average vectors
        self.a1 = [2.75, 1.5625, 2.34375, 2.20313, 2.71875, 2.90625, 1.9375,
                   2.32813, 2.51563, 1.9375, 2.42188, 2.40625, 2.42188, 1.92188, 2.20313]
        self.a2 = [2.37096774, 2.06451613, 1.421875, 1.734375, 2.109375, 1.953125,
                   2.21875, 2.09375,1.625, 2.03125, 2.21875, 2.25, 2.109375, 1.71875, 2.09375]
        self.b1 = [3.0625, 2.359375, 2.34375, 2.171875, 1.71875, 3.046875, 2.546875, 2.0625,
                   2.09375, 2.828125, 2.40625, 2.625, 2.328125, 2.734375]
        self.b2 = [2.203125, 2.890625, 2.09375, 1.46875, 1.890625, 2.375, 1.828125,
                   2.03125, 1.796875, 2.3125, 2.421875, 2.21875, 2.859375, 2.078125, 2.375]
        #Average vectors with only peer answers
        self.a1_peer = [3.25, 1.0625, 2.34375, 2.40625, 2.84375, 3.09375, 2.34375, 2.03125,
                        2.1875, 2.34375, 1.6875, 2.28125, 2.1875, 2.3125, 2.375]
        self.a2_peer = [2.6875, 2.75, 1.09375, 2.15625, 2.5625, 2.4375, 2.09375, 1.96875,
                        2.125, 1.5, 2.21875, 2.09375, 2.90625, 2.1875, 2.875]
        self.b1_peer = [3.125, 2.46875, 2.9375, 2, 1.5625, 3.53125, 3.5625, 2.0625,
                        2.40625, 3.75, 2.46875, 2.71875, 2.5, 2.5625]
        self.b2_peer = [2.15625, 3.09375, 1.90625, 1.71875, 1.46875, 2.4375, 2.4375,
                        2.15625, 2.125, 2.3125, 3.3125, 2, 3, 1.96875, 1.6875]
        #Average vectors with only selected questions
        self.a1_selected = [2.35, 1.4, 1.75, 1.55, 2.35, 2.55, 1.65, 2.15, 2.55, 1.95,
                            1.85, 2.45, 2.1, 1.35, 2.35]
        self.a2_selected = [2.05, 2, 1.25, 1.6, 2.1, 1.65, 2.25, 1.9, 1.25, 1.65, 2.2,
                            1.75, 2, 1.6, 1.85]
        self.b1_selected = [2.45 ,1.75 ,1.65 ,1.9 ,1.15 ,2.55 ,2.3 ,1.7 ,1.55 ,2.5 ,1.7 ,
                            2.25 ,2.15 ,2.55]
        self.b2_selected = [1.95, 2.8, 1.9, 1.3, 1.55, 2, 1.6, 2, 1.5, 2.05, 2.35, 1.3, 2.15,
                            2.05, 2.2]
        #Average vectors with only peer answers from selected questions
        self.a1_selected_peer = [2.7, 1, 2.1, 2.1, 2.1, 2.8, 2.2, 1.7, 2.4, 2.3, 1, 2, 2, 1.7, 2.6]
        self.a2_selected_peer = [2.1, 2.8, 1, 2.2, 2.4, 2.2, 2, 1.6, 1.5, 1, 2, 1.6, 2.7, 2, 2.6]
        self.b1_selected_peer = [2.5, 2, 2.3, 1.4, 1.1, 2.9, 3.6, 1.7, 2, 3.9, 1.5, 2.2, 2.2, 2.6]
        self.b2_selected_peer = [2, 3.4, 1.5, 1.4, 1, 2, 2.2, 2.2, 2, 2.1, 3.5, 1.3, 2.1, 2, 1.6]

        # dictionary to iterate in all the combination we need for the linear regressions
        self.features_list = ["features A1A2", "features B1B2", "features ALL"]
        self.a1a2dictionary = {"a1+a2":  self.a1 +  self.a2, "a1_peer+a2_peer":  self.a1_peer +  self.a2_peer,
                          "a1_selected+a2_selected":  self.a1_selected +  self.a2_selected,
                          "a1_selected_peer+a2_selected_peer":  self.a1_selected_peer +  self.a2_selected_peer}
        self.b1b2dictionary = {"b1+b2":  self.b1 +  self.b2, "b1_peer+b2_peer":  self.b1_peer +  self.b2_peer,
                          "b1_selected+b2_selected":  self.b1_selected +  self.b2_selected,
                          "b1_selected_peer+b2_selected_peer":  self.b1_selected_peer +  self.b2_selected_peer}
        self.alldictionary = {"all":  self.a1_selected_peer +  self.a2_selected_peer +  self.b1_selected_peer +  self.b2_selected_peer}


    def start(self):
        """
        This function upload the dataset and calls the linear regression function
        :return: the result of the method 'linear_comparison' between all initialized vectors and their datasets
        """
        #upload the dataset
        for feature in self.features_list:
            try:
                path = os.getcwd() + "/Analysis dataset/Features/" + feature + ".xlsx"
            except FileNotFoundError as e:
                print("Attenzione: la cartella 'Analysis dataset' non esiste ancora\n")
            else:
                path = os.getcwd() + "/Analysis dataset/Features/" + feature + ".xlsx"
                workbook = load_workbook(filename=path)
                sheet = workbook.active
                n_coloumn = sheet.max_column
                if feature == "features A1A2":
                    for key, vector in self.a1a2dictionary.items():
                        linear_comparison(sheet, vector, feature, key, n_coloumn)
                elif feature == "features B1B2":
                    for key, vector in self.b1b2dictionary.items():
                        linear_comparison(sheet, vector, feature, key, n_coloumn)
                else:
                    for key, vector in self.alldictionary.items():
                        linear_comparison(sheet, vector, feature, key, n_coloumn)



