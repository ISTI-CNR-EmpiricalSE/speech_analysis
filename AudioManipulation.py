#diarization
from pyannote.audio import Pipeline
#to work with the file system
import os
#for audio segmentation
from pydub import AudioSegment
#graphic visualization
from pyannote.core import Annotation, Segment
#to save and print the diarization and features images
from matplotlib import pyplot as plt
from pyannote.core import notebook
#to calculate the audio features
import librosa.display
import numpy as np
#to copy the file in the current folder
import shutil
#to build dataframe
import pandas as pd
#to save in excel dataset
from openpyxl import load_workbook

def librosa_average(librosa_matrix):
    """
    :param librosa_matrix: a matrix that contain one feature extracted from the file audio
    :return: a vector, in which every element represents the average of the elements contained in the relative vector within 'librosa_matrix'
    """
    shape = librosa_matrix.shape
    average_vector = []
    for i in range(0, shape[0]):
        average_vector.append(librosa_matrix[i].mean())
    return average_vector

def append_to_dictionary( features_name, data, dictionary):
    """
    This function append in a dictionary in which every pair <key,value> represents the i-th name+index of an element of a vector of features, and its value
    :param features_name: the name of the feature vector
    :param data: the feature vector
    :param dictionary: the dictionary in which to append the result
    :return:
    """
    for i in range (0, len(data)):
        string=features_name+"_"+str(i)
        list=[data[i]]
        dictionary[string]=list

class AudioManipulation:

    not_yet_calculated = []
    all_calculated=False
    audio_path = ""
    librosa_average_dictionary={}
    time_speakers=[]
    index=[]
    is_copied=False

    #features to be calculated
    features = {'stft_chromagram': np.array([])}
    features['constant_q_chromagram'] = np.array([])
    features['cens_chromagram'] = np.array([])
    features['melspectrogram'] = np.array([])
    features['mfcc'] = np.array([])
    features['rms'] = np.array([])
    features['spectral_centroid'] = np.array([])
    features['spectral_bandwidth'] = np.array([])
    features['spectral_contrast'] = np.array([])
    features['spectral_flatness'] = np.array([])
    features['spectral_rolloff'] = np.array([])
    features['polifeatures0'] = np.array([])
    features['polifeatures1'] = np.array([])
    features['polifeatures2'] = np.array([])
    features['tonnetz'] = np.array([])
    features['zero_crossing_rate'] = np.array([])
    features['time_speakers'] = np.array([])

    #flag to control if the features was yet calculated
    features_calculated = {'stft_chromagram': False}
    features_calculated['constant_q_chromagram'] = False
    features_calculated['cens_chromagram'] = False
    features_calculated['melspectrogram'] = False
    features_calculated['mfcc'] = False
    features_calculated['rms'] = False
    features_calculated['spectral_centroid'] = False
    features_calculated['spectral_bandwidth'] = False
    features_calculated['spectral_contrast'] = False
    features_calculated['spectral_flatness'] = False
    features_calculated['spectral_rolloff'] = False
    features_calculated['polifeatures'] = False
    features_calculated['tonnetz'] = False
    features_calculated['zero_crossing_rate'] = False
    features_calculated['diarization'] = False


    def __init__(self, audio_path):
        """
        Initialize all the parameters
        :return:
        """

        # extrapolate relative path if necessary
        if "/" in audio_path:
            split_path=audio_path.split('/')
            if not os.path.isfile(split_path[-1]):
                # copy temporarily to current folder
                shutil.copyfile(audio_path, split_path[-1])
            self.audio_path= split_path[-1]
            self.is_copied = True
        else:
            self.audio_path=audio_path
            self.is_copied = False

        #Initialize all the parameters
        for key in self.features:
            self.features[key] = np.array([])

        for key in self.features_calculated:
            self.features_calculated[key]=False

        self.not_yet_calculated = []
        self.all_calculated = False
        self.librosa_average_dictionary = {'name': self.audio_path}
        self.time_speakers = []



    def printime(self):
        """
        :return: a string as the path of the audio
        """
        print(self.audio_path)

    def get_features(self):
        """
        Every iteration of the endless while of this function extract a single feature from the audio file, saving the data
        in the relative parameters of the class, and the image in a directory.
        Since some of the commands work by concatenating others, a prioritized list called 'not_yet_calculated' was defined,
        which represents the list of the simple commands that compose it. If this structure is not empty, every iteration
        makes a pop on this list.
        :return:
        """
        # build the directory that will contain the images
        path = os.getcwd() + "/Visualization/"
        if not (os.path.isdir(path)):
            os.mkdir(path)
        path=os.getcwd() + "/Visualization/"+self.audio_path
        if not (os.path.isdir(path)):
            os.mkdir(path)

        #Analize the audio track
        y, sr = librosa.load(self.audio_path)
        # Separate harmonics and percussives into two waveforms
        y_harmonic, y_percussive = librosa.effects.hpss(y)
        S = np.abs(librosa.stft(y))

        still_features=True
        while still_features==True or len(self.not_yet_calculated)!=0:
            if len(self.not_yet_calculated)!=0:
                 feature_comand=self.not_yet_calculated[0]
                 self.not_yet_calculated.remove(self.not_yet_calculated[0])
            else:
                feature_comand = input("Digita 'help_features' per scegliere quale feature vuoi calcolare o 'stop_features' per smettere di calcolare feature\n")

            if (feature_comand=='stop_features'):
                if self.is_copied == True:
                    os.remove(self.audio_path)
                still_features=False

            elif(feature_comand=='help_features'):
                print(" 'stop_features': interrompe il calcolo delle features\n"
                      " 'diarization': effettua la diarizzazione dei file, esportandone anche una rappresentazione grafica e calcola il tempo totale dei parlanti, \n"
                      " 'stft_chromagram': calcola stft_chromagram\n"
                      " 'constant_q_chromagram': calcola constant_q_chromagram\n"
                      " 'cens_chromagram': calcola cens_chromagram\n"
                      " 'melspectrogram': calcola melspectrogram\n"
                      " 'mfcc': calcola mfcc\n"
                      " 'rms': calcola rms \n"
                      " 'spectral_centroid': calcola spectral_centroid\n"
                      " 'spectral_bandwidth': calcola spectral_bandwidth\n"
                      " 'spectral_contrast': calcola spectral_contrast\n"
                      " 'spectral_flatness': calcola spectral_flatness\n"
                      " 'spectral_rolloff': calcola spectral_rolloff\n"
                      " 'polifeatures': calcolo polifeatures\n"
                      " 'tonnetz': calcolo tonnetz\n"
                      " 'zero_crossing_rate': calcola zero_crossing_rate\n"
                      " 'all': calcola tutte le feature rimanenti\n"
                      " 'average': calcola un vettore concatenando le medie delle singole feature e lo salva un foglio excell")

            elif(feature_comand=='diarization'):
                #build the dyrectory who will contain the diarization
                path = os.getcwd() + "/Speakers/"
                if not (os.path.isdir(path)):
                    os.mkdir(path)

                # apply pretrained pipeline
                pipeline = Pipeline.from_pretrained("pyannote/speaker-diarization")
                diarization = pipeline(self.audio_path)

                # initialize variables to create the segment and the image from the diarization
                sound = AudioSegment.from_wav(self.audio_path)
                annotation = Annotation()

                # print the result
                counter_clips = 0  # counting the segment number
                path = os.getcwd() + "/Speakers/"+self.audio_path  # folder that will contain the speaker's folders
                if not (os.path.isdir(path)):
                    os.mkdir(path)
                for turn, _, speaker in diarization.itertracks(yield_label=True):  # print the diarization
                    #print(f"start={turn.start:.1f}s stop={turn.end:.1f}s speaker_{speaker}")
                    # visualization
                    annotation[Segment(turn.start, turn.end)] = speaker
                    # saving audio clips
                    start_clip = turn.start * 1000
                    end_clip = turn.end * 1000
                    #build the audio clip
                    output = sound[start_clip: end_clip]
                    string_counter_clips = str(counter_clips)
                    # create the direcory that will contain the speakers
                    path = os.getcwd() + "/Speakers/"+self.audio_path+"/"+speaker
                    if not (os.path.isdir(path)):
                        os.mkdir(path)
                    path = path + "/clip." + string_counter_clips + ".wav"
                    output.export(path, format="wav")
                    counter_clips = counter_clips + 1

                # save and print the diarization image
                plt.cla()
                plt.clf()
                #notebook.width = 10
                #plt.rcParams['figure.figsize'] = (notebook.width, 2)
                notebook.plot_annotation(annotation, legend=True, time=True)
                plt.savefig(os.getcwd() +"/Speakers/"+self.audio_path+"/diarization.jpg")
                plt.show()

                # counting speakers time
                path = os.getcwd() + "/Speakers/"+self.audio_path
                # Scan the directory with an os.DirEntry objects
                iterator = os.scandir(path)
                for entry in iterator:
                    if entry.is_dir():
                        path = os.getcwd() + "/Speakers/"+self.audio_path+"/"+entry.name
                        sub_iterator = os.scandir(path)
                        # scan speakers directories
                        time_count = 0
                        for sub_entry in sub_iterator:
                            if sub_entry.is_file():
                                # counting speakers time
                                file_path = path + "/" + sub_entry.name
                                audio = AudioSegment.from_file(file_path)
                                time_count = time_count + audio.duration_seconds
                        self.time_speakers.append(time_count)
                #considering only 2 speakers, the ones with higher time
                templist=[max(self.time_speakers)]
                self.time_speakers.remove(max(self.time_speakers))
                templist.append(max(self.time_speakers, default=0))
                self.features['time_speakers']= templist
                self.features_calculated['diarization'] = True

                print(self.audio_path,"diarization: Calcolo effettuato")

            elif (feature_comand=='stft_chromagram'):
                self.features['stft_chromagram'] = librosa.feature.chroma_stft(y=y_harmonic,sr=sr)
                self.features_calculated['stft_chromagram'] = True

                # visualization
                fig, ax = plt.subplots(nrows=2, sharex=True)
                img = librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max), y_axis='log', x_axis='time', ax=ax[0])
                fig.colorbar(img, ax=[ax[0]])
                ax[0].label_outer()
                img = librosa.display.specshow( self.features['stft_chromagram'], y_axis='chroma', x_axis='time', ax=ax[1])
                fig.colorbar(img, ax=[ax[1]])
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/stft_chromagram.jpg')

                print(self.audio_path,"stft_chromagram: Calcolo effettuato\n")

            elif (feature_comand=='constant_q_chromagram'):
                self.features['constant_q_chromagram'] = librosa.feature.chroma_cqt(y=y_harmonic, sr=sr)
                self.features_calculated['constant_q_chromagram'] = True

                #visualitazion
                fig, ax = plt.subplots()
                img = librosa.display.specshow(self.features['constant_q_chromagram'], y_axis='chroma', x_axis='time', ax=ax)
                ax.set(title='chroma_cqt')
                ax.label_outer()
                fig.colorbar(img, ax=ax)
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/constant_q_chromagram.jpg')

                print(self.audio_path,"constant_q_chromagram: Calcolo effettuato\n")

            elif(feature_comand== 'cens_chromagram'):
                self.features['cens_chromagram'] = librosa.feature.chroma_cens(y=y_harmonic, sr=sr)
                self.features_calculated['cens_chromagram'] = True

                #visualization
                fig, ax = plt.subplots()
                img = librosa.display.specshow( self.features['cens_chromagram'], y_axis='chroma', x_axis='time', ax=ax)
                ax.set(title='chroma_cens')
                fig.colorbar(img, ax=ax)
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/cens_chromagram.jpg')

                print(self.audio_path,"cens_chromagram: Calcolo effettuato\n")

            elif(feature_comand== 'melspectrogram'):
                self.features['melspectrogram'] = librosa.feature.melspectrogram(y=y_harmonic, sr=sr)
                self.features_calculated['melspectrogram'] = True
                
                #visualization
                fig, ax = plt.subplots()
                S_dB = librosa.power_to_db(S, ref=np.max)
                img = librosa.display.specshow(S_dB, x_axis='time', y_axis='mel', sr=sr, fmax=8000, ax=ax)
                fig.colorbar(img, ax=ax, format='%+2.0f dB')
                ax.set(title='Mel-frequency spectrogram')
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/melspectrogram.jpg')
                
                print(self.audio_path,"melspectrogram: Calcolo effettuato\n")

            elif(feature_comand=='mfcc'):
                self.features['mfcc'] = librosa.feature.mfcc(y=y_harmonic, sr=sr)
                self.features_calculated['mfcc'] = True
                
                #visualization
                fig, ax = plt.subplots()
                img = librosa.display.specshow( self.features['mfcc'], x_axis='time', ax=ax)
                fig.colorbar(img, ax=ax)
                ax.set(title='MFCC')
                ax.label_outer()
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/mfcc.jpg')

                print(self.audio_path,"mfcc: Calcolo effettuato\n")

            elif(feature_comand=='rms'):
                self.features['rms'] = librosa.feature.rms(y=y_harmonic)
                self.features_calculated['rms'] = True
                
                #visualization
                fig, ax = plt.subplots(nrows=2, sharex=True)
                times = librosa.times_like(self.features['rms'])
                ax[0].semilogy(times,  self.features['rms'][0], label='RMS Energy')
                ax[0].set(xticks=[])
                ax[0].legend()
                ax[0].label_outer()
                librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max), y_axis='log', x_axis='time', ax=ax[1])
                ax[1].set(title='log Power spectrogram')
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/rms.jpg')
                
                print(self.audio_path,"rms: Calcolo effettuato\n")

            elif(feature_comand=='spectral_centroid'):
                self.features['spectral_centroid'] = librosa.feature.spectral_centroid(y=y_harmonic, sr=sr)
                self.features_calculated['spectral_centroid'] = True
                
                #visualization
                times = librosa.times_like( self.features['spectral_centroid'])
                fig, ax = plt.subplots()
                librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max), y_axis='log', x_axis='time', ax=ax)
                ax.plot(times,  self.features['spectral_centroid'].T, label='Spectral centroid', color='w')
                ax.legend(loc='upper right')
                ax.set(title='log Power spectrogram')
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/spectral_centroid.jpg')
                
                print(self.audio_path,"spectral_centroid: Calcolo effettuato\n")

            elif(feature_comand=='spectral_bandwidth'):
                self.features['spectral_bandwidth'] = librosa.feature.spectral_bandwidth(y=y_harmonic, sr=sr)
                self.features_calculated['spectral_bandwidth'] = True
                
                #visualization
                fig, ax = plt.subplots(nrows=2, sharex=True)
                times = librosa.times_like(self.features['spectral_bandwidth'])
                centroid = librosa.feature.spectral_centroid(S=S)
                ax[0].semilogy(times, self.features['spectral_bandwidth'][0], label='Spectral bandwidth')
                ax[0].set(ylabel='Hz', xticks=[], xlim=[times.min(), times.max()])
                ax[0].legend()
                ax[0].label_outer()
                librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max),
                                         y_axis='log', x_axis='time', ax=ax[1])
                ax[1].set(title='log Power spectrogram')
                ax[1].fill_between(times, np.maximum(0, centroid[0] - self.features['spectral_bandwidth'][0]),
                                   np.minimum(centroid[0] + self.features['spectral_bandwidth'][0], sr / 2),
                                   alpha=0.5, label='Centroid +- bandwidth')
                ax[1].plot(times, centroid[0], label='Spectral centroid', color='w')
                ax[1].legend(loc='lower right')
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/spectral_bandwidth.jpg')
                
                print(self.audio_path,"spectral_bandwidth: Calcolo effettuato\n")

            elif(feature_comand=='spectral_contrast'):
                self.features['spectral_contrast'] = librosa.feature.spectral_contrast(S=S, sr=sr)
                self.features_calculated['spectral_contrast'] = True
                
                #visualization
                fig, ax = plt.subplots(nrows=2, sharex=True)
                img1 = librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max), y_axis='log', x_axis='time', ax=ax[0])
                fig.colorbar(img1, ax=[ax[0]], format='%+2.0f dB')
                ax[0].set(title='Power spectrogram')
                ax[0].label_outer()
                img2 = librosa.display.specshow(self.features['spectral_contrast'], x_axis='time', ax=ax[1])
                fig.colorbar(img2, ax=[ax[1]])
                ax[1].set(ylabel='Frequency bands', title='Spectral contrast')
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/spectral_contrast.jpg')
                
                print(self.audio_path,"spectral_contrast: Calcolo effettuato\n")

            elif(feature_comand=='spectral_flatness'):
                self.features['spectral_flatness'] = librosa.feature.spectral_flatness(y=y_harmonic)
                self.features_calculated['spectral_flatness'] = True
                #no visualization
                print(self.audio_path,"spectral_flatness: Calcolo effettuato\n")

            elif(feature_comand== 'spectral_rolloff'):
                self.rolloff_min = librosa.feature.spectral_rolloff(y=y, sr=sr, roll_percent=0.01)
                self.features['spectral_rolloff'] = librosa.feature.spectral_rolloff(y=y_harmonic, sr=sr)
                self.features_calculated['spectral_rolloff'] = True
                
                #visualization
                fig, ax = plt.subplots()
                librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max), y_axis='log', x_axis='time', ax=ax)
                ax.plot(librosa.times_like(self.features['spectral_rolloff']), self.features['spectral_rolloff'][0], label='Roll-off frequency (0.85)')
                ax.plot(librosa.times_like(self.features['spectral_rolloff']), self.rolloff_min[0], color='w',
                        label='Roll-off frequency (0.01)')
                ax.legend(loc='lower right')
                ax.set(title='log Power spectrogram')
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/spectral_rolloff.jpg')
                
                print(self.audio_path,"spectral_rolloff: Calcolo effettuato\n")

            elif(feature_comand=='polifeatures'):
                self.features['polifeatures0'] = librosa.feature.poly_features(S=S,order=0)  # Fit a degree-0 polynomial (constant) to each frame
                self.features['polifeatures1'] = librosa.feature.poly_features(S=S, order=1)  # Fit a linear polynomial to each frame
                self.features['polifeatures2'] = librosa.feature.poly_features(S=S, order=2)  # Fit a quadratic to each frame
                self.features_calculated['polifeatures'] = True

                #visualization
                fig, ax = plt.subplots(nrows=4, sharex=True, figsize=(8, 8))
                times = librosa.times_like(self.features['polifeatures0'])
                ax[0].plot(times, self.features['polifeatures0'][0], label='order=0', alpha=0.8)
                ax[0].plot(times, self.features['polifeatures1'][1], label='order=1', alpha=0.8)
                ax[0].plot(times, self.features['polifeatures2'][2], label='order=2', alpha=0.8)
                ax[0].legend()
                ax[0].label_outer()
                ax[0].set(ylabel='Constant term ')
                ax[1].plot(times, self.features['polifeatures1'][0], label='order=1', alpha=0.8)
                ax[1].plot(times, self.features['polifeatures2'][1], label='order=2', alpha=0.8)
                ax[1].set(ylabel='Linear term')
                ax[1].label_outer()
                ax[1].legend()
                ax[2].plot(times, self.features['polifeatures2'][0], label='order=2', alpha=0.8)
                ax[2].set(ylabel='Quadratic term')
                ax[2].legend()
                librosa.display.specshow(librosa.amplitude_to_db(S, ref=np.max), y_axis='log', x_axis='time', ax=ax[3])
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/polyfeatures.jpg')
                
                print(self.audio_path,"polifeatures: Calcolo effettuato\n")

            elif(feature_comand=='tonnetz'):
                self.features['tonnetz'] = librosa.feature.tonnetz(y=y, sr=sr)
                self.features_calculated['tonnetz'] = True
                
                #visualization
                fig, ax = plt.subplots(nrows=2, sharex=True)
                img1 = librosa.display.specshow(self.features['tonnetz'], y_axis='tonnetz', x_axis='time', ax=ax[0])
                ax[0].set(title='Tonal Centroids (Tonnetz)')
                ax[0].label_outer()
                img2 = librosa.display.specshow(librosa.feature.chroma_cqt(y=y, sr=sr), y_axis='chroma', x_axis='time', ax=ax[1])
                ax[1].set(title='Chroma')
                fig.colorbar(img1, ax=[ax[0]])
                fig.colorbar(img2, ax=[ax[1]])
                plt.savefig(os.getcwd() + "/Visualization/"+self.audio_path+'/tonnetz.jpg')
                
                print(self.audio_path,"tonnetz: Calcolo effettuato\n")

            elif (feature_comand== 'zero_crossing_rate'):
                self.features['zero_crossing_rate'] = librosa.feature.zero_crossing_rate(y)
                self.features_calculated['zero_crossing_rate'] = True
                print(self.audio_path,"zero_crossing_rate: Calcolo effettuato\n")


            elif (feature_comand == 'all'):
                """
                Extract all the remaining features
                """
                for key in self.features_calculated:
                    if (self.features_calculated[key] == False):
                        self.not_yet_calculated.insert(0, key)
                self.all_calculated=True

            elif(feature_comand == 'average'):
                """
                Build the excel file who will contain the average of all the features of the audio file
                """
                #If all the feature are jet calculated
                if self.all_calculated==True:
                    for feature in self.features:
                        if not feature=='time_speakers':
                            append_to_dictionary(feature,librosa_average(self.features[feature]), self.librosa_average_dictionary)
                    append_to_dictionary('time_speakers', self.features['time_speakers'], self.librosa_average_dictionary)
                    #Building a dataframe from the dictionary
                    librosa_average_dataframe = pd.DataFrame(data=self.librosa_average_dictionary)
                    # saving the dataframe on a xlsx file
                    path = os.getcwd() + "/Excel features/"
                    if not (os.path.isdir(path)):
                        os.mkdir(path)
                    #if the excel file already exists, it only adds a line to it
                    if os.path.isfile(os.getcwd() + "/Excel features/features.xlsx"):
                        book = load_workbook(os.getcwd() + "/Excel features/features.xlsx")
                        writer = pd.ExcelWriter(os.getcwd() + "/Excel features/features.xlsx", engine='openpyxl')
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        for sheetname in writer.sheets:
                            librosa_average_dataframe.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                                           index=False, header=False)
                        writer.save()
                    else:
                        librosa_average_dataframe.to_excel( os.getcwd() + "/Excel features/features.xlsx", index=False)

                    print("Dataframe esportato su file excel corettamente\n")

                else:
                    print("Esegui prima il comando 'all' ")

            else:
                print("La feature", feature_comand, "non esiste, ritenta")

    def get_average_vector(self):
        """
        Automatically extract and export all the features
        """
        self.not_yet_calculated.append('all')
        self.not_yet_calculated.append('average')
        self.not_yet_calculated.append('stop_features')
        self.get_features()



